from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
from django.db import transaction
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.contrib import messages

from .models import Product, Category, Manufacturer, Cart, CartItem, Order, OrderItem


def home_view(request):
    return render(request, 'shop/home.html')


def author_view(request):
    return render(request, 'shop/author.html')


def about_view(request):
    return render(request, 'shop/about.html')


def product_list(request):
    products = Product.objects.all()
    categories = Category.objects.all()
    manufacturers = Manufacturer.objects.all()

    category_id = request.GET.get('category')
    manufacturer_id = request.GET.get('manufacturer')
    search_query = request.GET.get('search')

    if category_id:
        products = products.filter(category_id=category_id)

    if manufacturer_id:
        products = products.filter(manufacturer_id=manufacturer_id)

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    context = {
        'products': products,
        'categories': categories,
        'manufacturers': manufacturers
    }

    return render(request, 'shop/product_list.html', context)


def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    return render(request, 'shop/product_detail.html', {'product': product})


@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    cart, created = Cart.objects.get_or_create(user=request.user)

    cart_item, created = CartItem.objects.get_or_create(
        cart=cart,
        product=product,
        defaults={'quantity': 1}
    )

    if not created:
        if cart_item.quantity + 1 <= product.stock_quantity:
            cart_item.quantity += 1
            cart_item.save()
        else:
            messages.error(request, "Недостаточно товара на складе.")

    return redirect('cart')


@login_required
def update_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)

    if request.method == 'POST':
        quantity = int(request.POST.get('quantity'))

        if quantity <= cart_item.product.stock_quantity:
            cart_item.quantity = quantity
            cart_item.save()
        else:
            messages.error(request, "Количество превышает остаток на складе.")

    return redirect('cart')


@login_required
def remove_from_cart(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    cart_item.delete()
    return redirect('cart')


@login_required
def cart_view(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    items = cart.items.all()
    total_price = cart.total_price()

    return render(request, 'shop/cart.html', {
        'cart': cart,
        'items': items,
        'total_price': total_price
    })

def build_receipt(order):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Чек"
    
    # Заголовок
    sheet["A1"] = f"Чек по заказу #{order.id}"
    sheet["A1"].font = Font(bold=True, size=14)
    
    # Информация о заказе
    sheet["A3"] = "Покупатель"
    sheet["B3"] = order.user.username
    sheet["A4"] = "Email"
    sheet["B4"] = order.email
    sheet["A5"] = "Адрес доставки"
    sheet["B5"] = order.shipping_address
    sheet["A6"] = "Дата заказа"
    sheet["B6"] = order.created_at.strftime("%d.%m.%Y %H:%M")
    
    # Товары
    headers = ["Товар", "Количество", "Цена", "Сумма"]
    sheet.append([])
    sheet.append(headers)
    for cell in sheet[8]:
        cell.font = Font(bold=True)
    
    for item in order.items.all():
        sheet.append([
            item.product_name,
            item.quantity,
            float(item.price),
            float(item.total_price())
        ])
    
    total_row = sheet.max_row + 1
    sheet.cell(total_row, 3, "Итого")
    sheet.cell(total_row, 4, float(order.total_price))
    sheet.cell(total_row, 3).font = Font(bold=True)
    sheet.cell(total_row, 4).font = Font(bold=True)
    
    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 15
    
    receipt = BytesIO()
    workbook.save(receipt)
    receipt.seek(0)
    return receipt


@login_required
def checkout(request):
    cart, created = Cart.objects.get_or_create(user=request.user)
    items = list(cart.items.select_related("product"))
    
    if not items:
        messages.error(request, "Корзина пуста. Добавьте товары перед оформлением заказа.")
        return redirect("cart")
    
    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        shipping_address = request.POST.get("shipping_address", "").strip()
        
        if not email or not shipping_address:
            messages.error(request, "Введите email и адрес доставки.")
            return redirect("checkout")
        
        with transaction.atomic():
            total_price = cart.total_price()
            order = Order.objects.create(
                user=request.user,
                email=email,
                shipping_address=shipping_address,
                total_price=total_price
            )
            
            for item in items:
                product = item.product
                
                if item.quantity > product.stock_quantity:
                    messages.error(
                        request,
                        f"Недостаточно товара '{product.name}' на складе."
                    )
                    order.delete()
                    return redirect("cart")
                
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    product_name=product.name,
                    price=product.price,
                    quantity=item.quantity
                )
                
                product.stock_quantity -= item.quantity
                product.save(update_fields=["stock_quantity"])
            
            cart.items.all().delete()
            
            
            messages.success(request, f"Заказ #{order.id} оформлен успешно!")
            return redirect("product_list")
    
    return render(request, "shop/checkout.html", {
        "items": items,
        "total_price": cart.total_price(),
        "default_email": request.user.email
    })